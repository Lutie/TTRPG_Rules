#!/usr/bin/env python3
"""
Parse Liaisons (L), Annexes (A) et Formes (F) depuis le xlsx magies.
Output: tools/mots-pouvoir.json

Règles de parsing :
- Row 1 = header → ignoré (min_row=2)
- Mot vide → stop (fin de la zone validée)
- Mot == "*" → skip (séparateur de section)
- Mot == "???" → skip (WIP sans nom)

Colonnes par tab :
  L: Mot | Type de Mot | Type de Sorts | Difficulté | Drain | Description | Conditions | Domaine
  A: Mot | Type de Mot | Type de Sorts | Difficulté | Drain | Description | Conditions | Domaine
  F: Mot | Type de Mot | Type de Forme | Limitations | Difficulté | Drain | Description | Particularités | Domaine
"""
import json
import openpyxl
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

XLSX = Path(__file__).parent / "src" / "Terre Natale - Aides de jeu _ Magies.xlsx"
OUT  = Path(__file__).parent / "mots-pouvoir.json"


def c(val):
    """Cell → stripped string, '' for None/None-string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s == "None" else s


def fmt_num(val):
    """'2.0' → '2', '-2.0' → '-2', 'X' → 'X', '' → ''."""
    if not val:
        return ""
    try:
        f = float(val)
        return str(int(f))
    except (ValueError, TypeError):
        return val


def parse_L(ws):
    """
    Liaisons — colonnes :
    0:Mot  1:TypeMot  2:TypeSorts  3:Diff  4:Drain  5:Desc  6:Cond  7:Domaine
    """
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = c(row[0] if len(row) > 0 else None)
        if not name:
            break                          # fin de la zone validée
        if name.startswith("*") or name.startswith("?"):
            continue                       # séparateur ou WIP sans nom

        out.append({
            "name":        name,
            "category":    "Liaison",
            "word_type":   "Liaison",
            "sort_type":   c(row[2] if len(row) > 2 else None),
            "difficulty":  fmt_num(c(row[3] if len(row) > 3 else None)),
            "drain":       fmt_num(c(row[4] if len(row) > 4 else None)),
            "description": c(row[5] if len(row) > 5 else None),
            "conditions":  c(row[6] if len(row) > 6 else None),
            "domain":      c(row[7] if len(row) > 7 else None),
            "mag_mod":     "",
            "sub_type":    "",
            "limitations": "",
        })
    return out


def parse_A(ws):
    """
    Annexes (Avancé) — même colonnes que L, mais conditions/domaine vides par conception.
    """
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = c(row[0] if len(row) > 0 else None)
        if not name:
            break
        if name.startswith("*") or name.startswith("?"):
            continue

        word_type = c(row[1] if len(row) > 1 else None) or "Avancé"

        out.append({
            "name":        name,
            "category":    "Annexe",
            "word_type":   word_type,
            "sort_type":   c(row[2] if len(row) > 2 else None),
            "difficulty":  fmt_num(c(row[3] if len(row) > 3 else None)),
            "drain":       fmt_num(c(row[4] if len(row) > 4 else None)),
            "description": c(row[5] if len(row) > 5 else None),
            "conditions":  "",
            "domain":      "",
            "mag_mod":     "",
            "sub_type":    "",
            "limitations": "",
        })
    return out


def parse_F(ws):
    """
    Formes — colonnes :
    0:Mot  1:TypeMot  2:TypeForme  3:Limitations  4:Diff  5:Drain  6:Desc  7:Particularités  8:Domaine
    """
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = c(row[0] if len(row) > 0 else None)
        if not name:
            break
        if name.startswith("*"):
            continue                       # séparateur de sous-type

        out.append({
            "name":        name,
            "category":    "Forme",
            "word_type":   "Forme",
            "sub_type":    c(row[2] if len(row) > 2 else None),   # Diffusion/Propagation/Structure
            "limitations": c(row[3] if len(row) > 3 else None),
            "difficulty":  fmt_num(c(row[4] if len(row) > 4 else None)),
            "drain":       fmt_num(c(row[5] if len(row) > 5 else None)),
            "description": c(row[6] if len(row) > 6 else None),
            "mag_mod":     c(row[7] if len(row) > 7 else None),   # Particularités
            "domain":      c(row[8] if len(row) > 8 else None),
            "sort_type":   "",
            "conditions":  "",
        })
    return out


OTHER_WORDS_OUT = Path(__file__).parent / "other_magic_words.json"

SHEET_CODES  = {"Liaison": "L", "Annexe": "A", "Forme": "F"}
GROUP_TYPES  = {"Liaison": "liaison", "Annexe": "annexe", "Forme": "forme"}


def ensure_dot(s):
    """Ajoute un point final si manquant."""
    s = s.strip()
    return s + "." if s and not s.endswith(".") else s


wb   = openpyxl.load_workbook(XLSX, data_only=True)
data = []
data.extend(parse_L(wb["L"]))
data.extend(parse_A(wb["A"]))
data.extend(parse_F(wb["F"]))

# ── mots-pouvoir.json (format étendu, pour les pages doc) ──────────────────
OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# ── other_magic_words.json (format compat pipeline_magie.py / app-sheet) ───
other = []
for i, w in enumerate(data, 1):
    other.append({
        "id":                 i,
        "sheet_code":         SHEET_CODES[w["category"]],
        "category":           w["category"],
        "group_type":         GROUP_TYPES[w["category"]],
        "name":               w["name"],
        "word_type":          w["word_type"],
        "difficulty":         w["difficulty"],
        "drain":              w["drain"],
        "description":        ensure_dot(w["description"]),
        "magnitude_modifiers": w["mag_mod"],
    })
OTHER_WORDS_OUT.write_text(json.dumps(other, ensure_ascii=False, indent=2), encoding="utf-8")

for cat in ("Liaison", "Annexe", "Forme"):
    n = sum(1 for w in data if w["category"] == cat)
    print(f"  {cat:10s}: {n:3d} entrées")
print(f"  {'Total':10s}: {len(data):3d} entrées")
print(f"Output : {OUT}")
print(f"Output : {OTHER_WORDS_OUT}")
