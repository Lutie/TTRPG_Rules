"""
Génère manoeuvres.json et prouesses.json depuis le fichier Excel des aides de jeu.
Usage : python3 tools/parse_manoeuvres.py
"""
import openpyxl
import json
import re
from pathlib import Path

XLSX = Path(__file__).parent / 'Terre Natale - Aides de jeu _ Manoeuvres & Prouesses.xlsx'
OUT_MAN = Path(__file__).parent / 'manoeuvres.json'
OUT_PRO = Path(__file__).parent / 'prouesses.json'


def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    s = re.sub(r'^❓\s*', '', s)
    s = s.strip()
    if s in ('-', 'n/a', 'N/A'):
        return ''
    return s


def parse_penalite(val):
    if val is None:
        return 0
    try:
        return int(float(str(val).replace('"', '').replace('+', '')))
    except:
        return 0


def parse_manoeuvres(wb):
    ws = wb['Manoeuvres']
    rows = list(ws.iter_rows(values_only=True))
    manoeuvres = []

    for row in rows[3:]:
        nom = row[0]
        typ = row[1]
        if not nom:
            continue        # ligne vide, on saute
        if not typ:
            break           # nom sans type = données en attente, on arrête

        manoeuvres.append({
            'nom':        clean(nom),
            'type':       clean(typ),
            'penalite':   parse_penalite(row[2]),
            'categorie':  clean(row[3]),
            'restrictions': {
                'att': bool(row[4]),
                'def': bool(row[5]),
                'tac': bool(row[6]),
            },
            'types_action': {
                'combat': bool(row[7]),
                'joute':  bool(row[8]),
                'mixte':  bool(row[9]),
                'autre':  bool(row[10]),
            },
            'description': clean(row[11]),
            'effets':      clean(row[12]),
            'modularite':  clean(row[13]),
            'notes':       clean(row[14]),
            'conditions':  clean(row[15]),
            'resume':      clean(row[16]),
        })

    return manoeuvres


def parse_prouesses(wb):
    ws = wb['Prouesses']
    rows = list(ws.iter_rows(values_only=True))
    prouesses = []

    for row in rows[3:]:
        nom_fluff = row[2]
        effets = row[4]
        if not nom_fluff or not effets:
            continue

        prouesses.append({
            'version':   clean(row[0]),
            'nom':       clean(row[1]) if row[1] else '',
            'nom_fluff': clean(nom_fluff),
            'type':      clean(row[3]),
            'effets':    clean(effets),
        })

    return prouesses


if __name__ == '__main__':
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    manoeuvres = parse_manoeuvres(wb)
    prouesses = parse_prouesses(wb)

    with open(OUT_MAN, 'w', encoding='utf-8') as f:
        json.dump(manoeuvres, f, ensure_ascii=False, indent=2)

    with open(OUT_PRO, 'w', encoding='utf-8') as f:
        json.dump(prouesses, f, ensure_ascii=False, indent=2)

    print(f'Manoeuvres : {len(manoeuvres)} → {OUT_MAN}')
    print(f'Prouesses  : {len(prouesses)} → {OUT_PRO}')
