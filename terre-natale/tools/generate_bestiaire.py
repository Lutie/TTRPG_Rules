#!/usr/bin/env python3
"""
Génère docs/tools/pnj/bestiaire_initial.json depuis tools/src/Bestiaire.xlsx.

Feuilles lues :
  Règnes   → regnes   (avec boosts/deboosts depuis texte attributs)
  Races    → races    (colonnes Boosts / Deboosts séparées)
  Ethnies  → ethnies  (colonnes Boosts / Deboosts séparées)
  Lignées  → lignees
  Actions  → actions  (bibliothèque)
  Effets (passifs)  → particularites (nouvelles uniquement)
  Effets (spéciaux) → particularites (nouvelles uniquement)

Usage : python3 tools/generate_bestiaire.py
"""

import json, re, unicodedata
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "tools/src/Bestiaire.xlsx"
OUT  = ROOT / "docs/tools/pnj/bestiaire_initial.json"

# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(text):
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    text = re.sub(r"[\s_-]+", "-", text.strip())
    return text

VALID_ATTRS = {"FOR","DEX","AGI","PER","CON","CHA","INT","RUS","SAG","VOL","MAG"}

def split_attrs(text):
    """'CON/FOR' ou 'FOR/DEX' → ['CON','FOR']  |  None/vide → []"""
    if not text:
        return []
    return [a.strip() for a in str(text).split("/") if a.strip() in VALID_ATTRS]

def parse_boosts_text(text):
    """'Boost FOR/VOL/CON, Deboost RUS/AGI/DEX' → (['FOR','VOL','CON'], ['RUS','AGI','DEX'])
    Pour les règnes dont les attributs sont encore en texte libre."""
    if not text or "Pas de modificateurs" in text:
        return [], []
    bm = re.search(r'Boost\s+([\w/]+)', text)
    dm = re.search(r'Deboost\s+([\w/]+)', text)
    boosts   = [a for a in (bm.group(1).split("/") if bm else []) if a in VALID_ATTRS]
    deboosts = [a for a in (dm.group(1).split("/") if dm else []) if a in VALID_ATTRS]
    return boosts, deboosts

def parse_attrs_sec(morphologie):
    """'STA 16, TAI 16, EGO 12, APP 8, CHN 8, EQU 12' → ajustements depuis base 10."""
    base = {"STA": 0, "TAI": 0, "EGO": 0, "APP": 0, "CHN": 0, "EQU": 0}
    if not morphologie:
        return base
    for m in re.finditer(r'(STA|TAI|EGO|APP|CHN|EQU)\s+(\d+)', str(morphologie)):
        base[m.group(1)] = int(m.group(2)) - 10
    return base

def part_text(pid, nom, description, type_part="passif"):
    return {
        "id": pid, "nom": nom, "type_part": type_part, "mode": "texte",
        "sous_effets": [{"nom": nom, "description": description}]
    }

def action_name(desc):
    """Génère un nom court depuis la description d'une action."""
    quoted = re.search(r'"([^"]{3,30})"', desc)
    if quoted:
        return quoted.group(1).capitalize()
    for pattern, name in [
        (r'illusion',                               "Illusions"),
        (r'clone',                                  "Clone"),
        (r"prend.*apparence",                       "Métamorphose"),
        (r"annule.*sort",                           "Contresort"),
        (r"projectile",                             "Multiprojectile"),
        (r"souffle.*énergie",                       "Souffle élémentaire"),
        (r"zone.*énergie",                          "Zone d'énergie"),
        (r"gaz toxique",                            "Gaz toxique"),
        (r"séisme",                                 "Séisme"),
        (r"vent.*repousse",                         "Bourrasque"),
        (r"avaler",                                 "Avalement"),
        (r"brouillard.*\{2x\}",                     "Brouillard"),
        # métamorphoses avant les patterns génériques
        (r"change.*loups",                          "Métamorphose (loups)"),
        (r"change.*brouillard",                     "Métamorphose (brouillard)"),
        (r"change.*nuée|chauve.*souris",            "Nuée (chauve-souris)"),
        (r"intangible",                             "Intangibilité"),
        (r"invisible",                              "Invisibilité"),
        (r"téléporte",                              "Téléportation"),
        (r"test de robustesse.*moitié|moitié.*PV",  "Mort instantanée"),
        (r"attire.*cibles",                         "Aspiration"),
        (r"repousse.*cibles",                       "Répulsion"),
        (r"copie.*action",                          "Imitation"),
        (r"vers.*corps",                            "Vers parasites"),
        (r"dégradation.*métal",                     "Dégradation du métal"),
        (r"enraciné",                               "Enracinement"),
    ]:
        if re.search(pattern, desc, re.IGNORECASE):
            return name
    # Fallback : premiers mots significatifs
    short = re.sub(r"^La créature\s+", "", desc, flags=re.IGNORECASE)
    return short[:40].rstrip(" ,") + "…"

# ── Particularités injectées manuellement (règnes/ethnies) ────────────────────

MANUAL_PARTICULARITES = [
    # Règnes
    part_text("qualite-objets-ameliorations", "Qualité des objets et améliorations",
              "La qualité des objets et des améliorations disponibles augmente de {x}."),
    part_text("styles-de-combat", "Styles de combat",
              "La créature maîtrise des styles de combat supplémentaires de niveau {x}."),
    part_text("recuperation-surnaturelle", "Récupération surnaturelle",
              "Récupération +{x} (toutes ressources)."),
    part_text("armes-rudimentaires", "Armes rudimentaires",
              "La créature utilise des armes rudimentaires de niveau {x}."),
    part_text("adrenaline-surnaturelle", "Adrénaline surnaturelle",
              "Adrénaline surnaturelle {x} : la créature entre en frénésie sous certaines conditions."),
    part_text("ferocite-competente", "Férocement compétent",
              "Compétence de base = 0, mais les attributs corporels remplacent le dé de compétence avec un bonus de {x}."),
    part_text("pattern-de-combat", "Pattern de combat",
              "La créature suit un schéma de combat instinctif de niveau {x}."),
    part_text("naturellement-competent", "Naturellement compétent",
              "La compétence est doublée (×2), mais chaque point de compétence réduit les attributs d'autant. Bonus {x}."),
    part_text("opposition-naturelle", "Opposition naturelle",
              "Opposition +{x}."),
    part_text("vitalite-surnaturelle", "Vitalité surnaturelle",
              "Vitalité surnaturelle {x} : PV et récupération physique fortement amplifiés."),
    part_text("maladie-physique-passive", "Maladie physique",
              "La créature est porteuse d'une maladie physique de catégorie {x}."),
    part_text("spiritualite-surnaturelle", "Spiritualité surnaturelle",
              "Spiritualité surnaturelle {x} : PS et résilience mentale fortement amplifiés."),
    part_text("maladie-mentale-passive", "Maladie mentale",
              "La créature est porteuse d'une maladie mentale de catégorie {x}."),
    part_text("immunite-speciale", "Immunité spéciale",
              "Immunité {x} : résistance majeure ou invulnérabilité, sauf condition particulière précisée."),
    part_text("implacable", "Implacable",
              "Implacable {x} : inflige des dégâts extrêmes, sauf condition particulière précisée."),
    part_text("garde-surnaturelle", "Garde surnaturelle",
              "Garde surnaturelle {x} : défenses physiques et mentales fortement amplifiées."),
    part_text("extremes-sauvegardes", "Extrêmes",
              "Extrêmes {x} : les sauvegardes les plus hautes reçoivent un bonus, les plus basses une pénalité."),
    part_text("chi-surnaturelle", "Chi surnaturelle",
              "Chi surnaturelle {x} : énergie et récupération de PK fortement amplifiées."),
    part_text("multiple-nature", "Multiple nature",
              "Multiple nature {x} : la créature est divisée en {x} points vitaux distincts."),
    part_text("imperieux", "Impérieux",
              "Impérieux {x} : endurance et résistance surnaturelles."),
    part_text("chroma", "Chroma",
              "Chroma {x} : affinité élémentaire renforcée selon la nature de la créature."),
    part_text("mana-surnaturelle", "Mana surnaturelle",
              "Mana surnaturelle {x} : PM et récupération magique fortement amplifiés."),
    part_text("magie-elementaire-passive", "Magie élémentaire",
              "Magie {x} : accès aux énergies élémentaires et sorts élémentaires de niveau {x}."),
    part_text("karma-surnaturelle", "Karma surnaturelle",
              "Karma surnaturelle {x} : Chance et Équilibre fortement amplifiés."),
    part_text("magie-divine-passive", "Magie divine / occulte",
              "Magie {x} : accès au divin et à l'occulte de niveau {x}."),
    part_text("solide", "Solide",
              "Solide {x} : les dégâts subis sont basés sur la dégradation de la créature."),
    # Ethnies
    part_text("creature-du-froid", "Créature du froid",
              "Créature du froid {x} : résistances et aptitudes liées aux environnements glaciaux."),
    part_text("odorat-sang", "Odorat spécialisé (sang)",
              "Odorat spécialisé : sang. Compétence spécialisée +{2x}."),
    {
        "id": "regeneration-sang-passive",
        "nom": "Régénération par le sang",
        "type_part": "passif", "mode": "composee",
        "declencheur_id": "lorsque-touchee-melee",
        "effets": [{"effet_id": "drain-ressource", "parametres": {"cible": "PV", "ressource": "PV"}}],
        "sous_effets": []
    },
    # Nouveaux passifs depuis Effets (passifs)
    part_text("defense-attaques", "Défenses contre les attaques",
              "Défenses contre les attaques +{x} (sauf si au sol)."),
    part_text("defense-tactiques", "Défenses contre les tactiques",
              "Défenses contre les tactiques +{x} (sauf si blessé)."),
]

NEW_REGLE = {
    "id": "sylvestre-naturel",
    "nom": "Naturel sylvestre",
    "description": "Tous les attributs sont défavorisés, mais la Chance et l'Équilibre reçoivent un double boost."
}

# ── Tables de correspondance règne → règles / particularités ──────────────────

REGNE_RULES = {
    "humanoide":   ["aucun-modificateur-regne"],
    "animal":      ["corps-favorise-esprit-defavorise"],
    "sylvestre":   ["sylvestre-naturel"],
    "necrophage":  ["esprit-nul"],
    "spectral":    ["corps-nul"],
    "maudit":      ["boost-general-chance-reduite"],
    "abomination": ["double-corps-double-esprit"],
    "chimerique":  ["double-corps-double-esprit"],
    "dragonoide":  ["affinite-elementaire"],
    "elementaire": ["affinite-elementaire"],
    "primordial":  ["boost-general-equilibre-reduit"],
    "artificiel":  ["artificiel-sans-esprit"],
}

REGNE_PARTS = {
    "humanoide":   ["qualite-objets-ameliorations", "styles-de-combat"],
    "ogroide":     ["recuperation-surnaturelle", "armes-rudimentaires"],
    "vermine":     ["adrenaline-surnaturelle", "armes-rudimentaires"],
    "animal":      ["ferocite-competente", "pattern-de-combat"],
    "sylvestre":   ["naturellement-competent", "opposition-naturelle"],
    "necrophage":  ["vitalite-surnaturelle", "maladie-physique-passive"],
    "spectral":    ["spiritualite-surnaturelle", "maladie-mentale-passive"],
    "maudit":      ["immunite-speciale", "implacable"],
    "abomination": ["garde-surnaturelle", "extremes-sauvegardes"],
    "chimerique":  ["chi-surnaturelle", "multiple-nature"],
    "dragonoide":  ["imperieux", "chroma"],
    "elementaire": ["mana-surnaturelle", "magie-elementaire-passive"],
    "primordial":  ["karma-surnaturelle", "magie-divine-passive"],
    "artificiel":  ["solide", "opposition-naturelle"],
}

RACE_PARTS = {
    "urside": ["protection", "absorption"],
}

ETHNIE_PARTS = {
    "blanc":      ["attrition", "creature-du-froid"],
    "rouge-sang": ["regeneration-sang-passive", "odorat-sang"],
    "venimeux":   ["morsure-venimeuse"],
}

# ── Chargement JSON existant ──────────────────────────────────────────────────

with open(OUT, encoding="utf-8") as f:
    data = json.load(f)
data["version"] = 2

# Règle Sylvestre
existing_rule_ids = {r["id"] for r in data.get("regles", [])}
if NEW_REGLE["id"] not in existing_rule_ids:
    data["regles"].append(NEW_REGLE)

# Particularités manuelles
existing_part_ids = {p["id"] for p in data.get("particularites", [])}
for p in MANUAL_PARTICULARITES:
    if p["id"] not in existing_part_ids:
        data["particularites"].append(p)
        existing_part_ids.add(p["id"])

# ── Lecture xlsx ──────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)

# ── Règnes ────────────────────────────────────────────────────────────────────

regnes = []
for row in list(wb["Règnes"].iter_rows(values_only=True))[1:]:
    nom = row[0]
    if not nom:
        continue
    rid = slugify(nom)
    boosts, deboosts = parse_boosts_text(row[1] or "")
    regnes.append({
        "id": rid, "nom": nom,
        "boosts": boosts, "deboosts": deboosts,
        "competences": [], "particularites": REGNE_PARTS.get(rid, []),
        "actions": [], "regles": REGNE_RULES.get(rid, []),
    })

# ── Races ─────────────────────────────────────────────────────────────────────

regne_by_nom = {r["nom"]: r["id"] for r in regnes}

races = []
for row in list(wb["Races"].iter_rows(values_only=True))[1:]:
    nom = row[0]
    if not nom or nom == "Universel":
        continue
    rid = slugify(nom)
    races.append({
        "id": rid, "nom": nom,
        "regne": regne_by_nom.get(row[1] or ""),
        "boosts": split_attrs(row[2]),
        "deboosts": split_attrs(row[3]),
        "competences": [], "particularites": RACE_PARTS.get(rid, []),
        "actions": [],
        "arme_nat": 0, "armure_nat": 0,
        "attrs_sec": parse_attrs_sec(row[4]),
    })

# ── Ethnies ───────────────────────────────────────────────────────────────────

ethnies = []
for row in list(wb["Ethnies"].iter_rows(values_only=True))[1:]:
    nom = row[0]
    if not nom:
        continue
    eid = slugify(nom)
    ethnies.append({
        "id": eid, "nom": nom,
        "race": None,
        "boosts": split_attrs(row[2]),
        "deboosts": split_attrs(row[3]),
        "competences": [], "particularites": ETHNIE_PARTS.get(eid, []),
        "actions": [],
    })

# ── Lignées ───────────────────────────────────────────────────────────────────

lignees = []
for row in list(wb["Lignées"].iter_rows(values_only=True))[1:]:
    nom = row[0]
    if not nom:
        continue
    lid = slugify(nom)
    boosts = [a.strip() for a in [str(row[1] or ""), str(row[2] or "")] if a.strip() in VALID_ATTRS]
    lignees.append({
        "id": lid, "nom": nom,
        "regne": None,
        "boosts": boosts, "deboosts": [],
        "competences": [], "particularites": [], "actions": [],
        "sauvegardes": {},
    })

# ── Actions (bibliothèque) ────────────────────────────────────────────────────
# On repart d'une liste vide à chaque run pour éviter les doublons

data["actions"] = []
existing_action_ids: set = set()
added_actions = 0
for row in list(wb["Actions"].iter_rows(values_only=True))[1:]:
    desc = row[1]
    if not desc or not str(desc).strip():
        continue
    desc = str(desc).strip()
    nom  = action_name(desc)
    aid  = slugify(nom)
    # Dédoublonnage par slug
    base_aid, n = aid, 1
    while aid in existing_action_ids:
        n += 1
        aid = f"{base_aid}-{n}"
    data["actions"].append({"id": aid, "nom": nom, "description": desc})
    existing_action_ids.add(aid)
    added_actions += 1

# ── Injection ─────────────────────────────────────────────────────────────────

data["regnes"]  = regnes
data["races"]   = races
data["ethnies"] = ethnies
data["lignees"] = lignees

# ── Écriture ──────────────────────────────────────────────────────────────────

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"✓ {OUT}")
print(f"  {len(regnes)} règnes · {len(races)} races · {len(ethnies)} ethnies · {len(lignees)} lignées")
print(f"  {len(data['particularites'])} particularités · {len(data['regles'])} règles · {len(data['actions'])} actions")
